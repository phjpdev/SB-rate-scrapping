import re
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

ua = UserAgent()
USER_AGENT = ua.random
DEFAULT_CHROMEDRIVER_PATH = r"C:\chromedriver\chromedriver.exe"

BASE_URL = 'https://www.sportsbet.com.au'
FILE_NAME = 'Race Meetings.xlsm'
target_column = 23
ALLOWED_MEETINGS = ['(VIC)', '(NSW)', '(QLD)', '(SA)', '(WA)', '(NT)', '(TAS)', '(ACT)', '(NZ)', '(NZL)']
FS = {}
SR = {}
NON_MEETING_SHEETS = {"Current R&S Settings", "Instructions", "Track Conditions"}

def setup_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-images")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    options.add_argument(f"--user-agent={USER_AGENT}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--ignore-certificate-errors")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    chromedriver_path = (
        os.getenv("CHROMEDRIVER_PATH")
        or (DEFAULT_CHROMEDRIVER_PATH if os.path.exists(DEFAULT_CHROMEDRIVER_PATH) else None)
    )
    service = Service(chromedriver_path) if chromedriver_path else Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(800)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver

def extract_sb_rating(driver, race_url, meeting_key: str):
    global SR

    parsed = _parse_schedule_href(race_url) if isinstance(race_url, str) else None
    race_no = parsed[1] if parsed else None

    driver.get(BASE_URL + race_url)
    wait = WebDriverWait(driver, 20)

    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "div[data-automation-id^='racecard-outcome-']")
        )
    )

    if race_no:
        print(f"=== {meeting_key} R{race_no} === {race_url}", flush=True)
    else:
        print(f"=== {meeting_key} === {race_url}", flush=True)

    # Speed optimization (from script-111.py):
    # Expand Form once and parse all runner shortforms in one pass.
    try:
        expand_btn = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "span[data-automation-id='racecard-expand-form']")
            )
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", expand_btn)
        time.sleep(0.15)
        driver.execute_script("arguments[0].click();", expand_btn)
    except Exception:
        # already expanded or not present
        pass

    try:
        wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div[data-automation-id^='shortform-']")
            )
        )
    except Exception:
        # If shortforms never appear, skip this race.
        return

    soup = BeautifulSoup(driver.page_source, "html.parser")
    shortforms = soup.select("div[data-automation-id^='shortform-']")

    for sf in shortforms:
        try:
            sf_id = sf.get("data-automation-id", "")
            m = re.search(r"shortform-(\d+)", sf_id)
            if not m:
                continue
            runner_id = m.group(1)

            racecard = soup.select_one(
                f"div[data-automation-id='racecard-outcome-{runner_id}']"
            )
            if not racecard:
                continue

            name_el = racecard.select_one(
                "div[data-automation-id='racecard-outcome-name'] span"
            )
            if not name_el:
                continue

            horse_name = re.sub(r"^\d+\.\s*", "", name_el.get_text(strip=True))

            sb_el = sf.select_one(
                "div[data-automation-id='shortform-SB Rating'] span:last-child"
            )
            if not sb_el:
                continue

            sb_rating = sb_el.get_text(strip=True)

            win_fixed = None
            win_el = racecard.select_one("[data-automation-id='racecard-outcome-0-L-price']")
            if win_el:
                t = win_el.get_text(" ", strip=True)
                if t and re.fullmatch(r"\d{1,3}\.\d{2}", t):
                    win_fixed = t

            SR.setdefault(meeting_key, {})
            SR[meeting_key].setdefault(horse_name, {})
            SR[meeting_key][horse_name]["sb_rating"] = sb_rating
            if win_fixed is not None:
                SR[meeting_key][horse_name]["win_fixed"] = win_fixed

            if win_fixed is not None:
                print(f"[OK] {horse_name} -> SB Rating {sb_rating} | Win Fixed {win_fixed}", flush=True)
            else:
                print(f"[OK] {horse_name} -> SB Rating {sb_rating}", flush=True)

        except Exception:
            continue

def disable_international_filter(driver):
    wait = WebDriverWait(driver, 20)

    try:
        # Detect ON state
        on_state = driver.find_elements(
            By.CSS_SELECTOR,
            "div[data-automation-id='filter-button-international-on']"
        )

        if not on_state:
            print("Int'l already OFF")
            return

        print("Int'l filter ON -> disabling", flush=True)

        # Click the LABEL (this is critical)
        label = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "label[for='ALL_RACING_PAGEINTERNATIONAL']")
            )
        )

        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", label
        )
        time.sleep(0.3)

        driver.execute_script("arguments[0].click();", label)

        # WAIT FOR OFF STATE
        wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div[data-automation-id='filter-button-international-off']")
            )
        )

        # WAIT FOR TABLE RE-DRAW (critical)
        wait.until(
            EC.staleness_of(
                driver.find_element(
                    By.CSS_SELECTOR,
                    "td[data-automation-id^='horse-racing-section-row-']"
                )
            )
        )

        print("Int'l filter OFF", flush=True)

    except Exception as e:
        print(f"Failed to disable Int'l filter ({type(e).__name__}): {e}", flush=True)

def _norm_text(s: str) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", (s or "").upper()).strip()

def _slugify(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")

def _extract_race_no(label: str, href: str | None = None) -> int | None:
    for text in (label or "", href or ""):
        m = re.search(r"\bR(?:ACE)?\s*(\d{1,2})\b", text, flags=re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                pass
        m = re.search(r"/race[-/](\d{1,2})\b", text, flags=re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                pass
    return None

def get_excel_targets(excel_file: str) -> dict[str, set[int]]:
    """
    Returns meeting -> set(race_numbers) derived from the workbook.
    Heuristic: each race block has a 'Horse' header in column D; we scan upward
    to find the race number in column A, and confirm at least 1 horse exists below.
    """
    wb = load_workbook(filename=excel_file, keep_vba=True, data_only=False)
    targets: dict[str, set[int]] = {}

    for ws in wb.worksheets:
        if ws.title in NON_MEETING_SHEETS:
            continue

        meeting_name = ws["G1"].value if isinstance(ws["G1"].value, str) and ws["G1"].value.strip() else ws.title
        meeting_name = str(meeting_name).strip()

        races: set[int] = set()
        max_row = min(ws.max_row or 0, 2000)

        for r in range(1, max_row + 1):
            d = ws.cell(r, 4).value  # column D
            if not (isinstance(d, str) and d.strip().upper() == "HORSE"):
                continue

            # find race number above
            race_no = None
            for rr in range(r, max(r - 20, 1) - 1, -1):
                a = ws.cell(rr, 1).value  # column A
                if isinstance(a, (int, float)) and int(a) == a and 1 <= int(a) <= 24:
                    race_no = int(a)
                    break

            if not race_no:
                continue

            # confirm there is at least one horse name below header
            has_horse = False
            for rr in range(r + 1, min(r + 30, max_row) + 1):
                v = ws.cell(rr, 4).value
                if isinstance(v, str) and v.strip().upper() == "HORSE":
                    break
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                has_horse = True
                break

            if has_horse:
                races.add(race_no)

        if races:
            targets[meeting_name] = races

    return targets

def _parse_schedule_href(href: str) -> tuple[str, int] | None:
    """
    Parses:
      /horse-racing/australia-nz/<meeting-slug>/race-<n>-<id>
    Returns (meeting_slug, race_no).
    """
    m = re.match(r"^/horse-racing/australia-nz/([^/]+)/race-(\d{1,2})-\d+", href, flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return (m.group(1).lower(), int(m.group(2)))
    except Exception:
        return None

def get_races(driver):
    excel_targets = get_excel_targets(FILE_NAME)
    if excel_targets:
        print(
            "Excel targets: "
            + ", ".join(f"{m} R{sorted(rs)}" for m, rs in excel_targets.items()),
            flush=True,
        )
        expected = sum(len(rs) for rs in excel_targets.values())
        print(f"Expected race count from Excel: {expected}", flush=True)
    else:
        print("Excel targets: none found (falling back to schedule filtering).", flush=True)

    driver.get(BASE_URL + "/racing-schedule")
    wait = WebDriverWait(driver, 20)
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "td[data-automation-id^='horse-racing-section-row-']")
        )
    )

    soup = BeautifulSoup(driver.page_source, "html.parser")
    race_links: list[str] = []
    seen: set[str] = set()

    if not excel_targets:
        for a in soup.select("a[href]"):
            href = a.get("href")
            if not isinstance(href, str) or not href.startswith("/"):
                continue
            if "/horse-racing/australia-nz/" not in href.lower():
                continue
            if "/race-" not in href.lower():
                continue
            if href in seen:
                continue
            seen.add(href)
            race_links.append(href)
        print(f"Filtered races (AU/NZ only): {len(race_links)}", flush=True)
        return race_links

    slug_to_meeting = { _slugify(m): m for m in excel_targets.keys() }
    available_meetings: set[str] = set()
    per_meeting_count: dict[str, int] = {m: 0 for m in excel_targets.keys()}

    for a in soup.select("a[href]"):
        href = a.get("href")
        if not isinstance(href, str) or not href.startswith("/"):
            continue

        parsed = _parse_schedule_href(href)
        if not parsed:
            continue

        meeting_slug, race_no = parsed
        available_meetings.add(meeting_slug)

        meeting_name = slug_to_meeting.get(meeting_slug)
        if not meeting_name:
            continue

        if href in seen:
            continue
        seen.add(href)
        race_links.append(href)
        per_meeting_count[meeting_name] = per_meeting_count.get(meeting_name, 0) + 1

    print(f"Filtered races (Excel only): {len(race_links)}", flush=True)
    print(
        "Races found per meeting: "
        + ", ".join(f"{m}={per_meeting_count.get(m, 0)}" for m in excel_targets.keys()),
        flush=True,
    )
    if race_links:
        # show first few selected races so it’s obvious we’re not scraping everything
        preview = []
        for href in race_links[:12]:
            parsed = _parse_schedule_href(href)
            if not parsed:
                preview.append(href)
                continue
            ms, rn = parsed
            preview.append(f"{ms} R{rn} {href}")
        print("Selected races (preview):", flush=True)
        for p in preview:
            print(f"- {p}", flush=True)
    if len(race_links) == 0:
        print(
            "No Excel meetings were found on the current Sportsbet schedule. "
            "Available meeting slugs on schedule include: "
            + ", ".join(sorted(list(available_meetings))[:40]),
            flush=True,
        )

    # Sort for nicer logs: meeting then race number.
    def _sort_key(h: str):
        parsed = _parse_schedule_href(h)
        if not parsed:
            return ("zzzz", 999, h)
        ms, rn = parsed
        return (ms, rn, h)

    race_links.sort(key=_sort_key)
    return race_links

def normalize_horse(name: str) -> str:
    return (
        name.strip()
        .upper()          # case-insensitive
        .replace(".", "") # remove dots like "5. "
    )


def save_sb_to_excel(excel_file, SR):
    workbook = load_workbook(filename=excel_file, keep_vba=True)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name in NON_MEETING_SHEETS:
            continue

        meeting_key = sheet["G1"].value if isinstance(sheet["G1"].value, str) and sheet["G1"].value.strip() else sheet_name
        meeting_key = str(meeting_key).strip()
        meeting_ratings = SR.get(meeting_key, {})
        if not meeting_ratings:
            continue

        # Speed: prebuild lookup map once per sheet
        norm_map = {normalize_horse(k): v for k, v in meeting_ratings.items()}

        for row in sheet.iter_rows(min_row=1):
            horse_cell = row[3]  # Column D
            if not horse_cell.value:
                continue

            excel_horse = normalize_horse(str(horse_cell.value))
            data = norm_map.get(excel_horse)
            if not data:
                continue

            sb_rating = data.get("sb_rating") if isinstance(data, dict) else data
            win_fixed = data.get("win_fixed") if isinstance(data, dict) else None

            if sb_rating is not None:
                sheet.cell(row=horse_cell.row, column=25, value=sb_rating)  # Y Sportsbet Rating
            if win_fixed is not None:
                sheet.cell(row=horse_cell.row, column=22, value=float(win_fixed))  # V Sportsbet Odds (Win Fixed)

            msg = f"Saved | {horse_cell.value}"
            if sb_rating is not None:
                msg += f" | SB {sb_rating}"
            if win_fixed is not None:
                msg += f" | Win {win_fixed}"
            print(msg, flush=True)

    try:
        workbook.save(excel_file)
        print(f"Workbook saved: {excel_file}", flush=True)
        return
    except PermissionError:
        # If Excel has the workbook open, OpenPyXL can't overwrite it.
        # Use Excel COM automation to write + Save() in-place.
        print(
            f"'{excel_file}' is in use (likely open in Excel). "
            "Attempting to save via Excel...",
            flush=True,
        )

    save_sb_to_excel_via_excel_com(excel_file, SR)


def save_sb_to_excel_via_excel_com(excel_file: str, SR: dict) -> None:
    try:
        import win32com.client  # type: ignore
        import pywintypes  # type: ignore
    except Exception as e:
        print(
            "Excel file is locked and Excel-automation is unavailable. "
            "Install it with: pip install pywin32",
            flush=True,
        )
        return

    excel_file_abs = str(Path(excel_file).resolve())
    target_name = Path(excel_file_abs).name.lower()

    xl = None
    wb = None

    try:
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            print(
                "Excel is not running. Close the workbook (or open Excel) and rerun.",
                flush=True,
            )
            return

        # Find already-open workbook IN THIS EXCEL INSTANCE.
        # IMPORTANT: do NOT open a second copy (can cause sharing violation).
        for w in xl.Workbooks:
            try:
                full = str(getattr(w, "FullName", "") or "")
                name = str(getattr(w, "Name", "") or "")
                if full and full.lower() == excel_file_abs.lower():
                    wb = w
                    break
                if name and name.lower() == target_name:
                    wb = w  # fallback match by name
            except Exception:
                continue

        if wb is None:
            print(
                f"Workbook not found among open Excel workbooks: {excel_file_abs}. "
                "Make sure THIS file is open (and only in one Excel instance), then rerun.",
                flush=True,
            )
            return

        try:
            if bool(getattr(wb, "ReadOnly", False)):
                print(
                    f"Workbook is open read-only in Excel, cannot save in-place: {excel_file_abs}",
                    flush=True,
                )
                return
        except Exception:
            pass

        # avoid prompts interrupting automation
        prev_alerts = xl.DisplayAlerts
        xl.DisplayAlerts = False

        try:
            for ws in wb.Worksheets:
                sheet_name = str(ws.Name)
                if sheet_name in NON_MEETING_SHEETS:
                    continue

                meeting_key = ws.Range("G1").Value
                meeting_key = str(meeting_key).strip() if meeting_key else sheet_name
                meeting_ratings = SR.get(meeting_key, {})
                if not meeting_ratings:
                    continue

                used = ws.UsedRange
                last_row = int(used.Row) + int(used.Rows.Count) - 1
                if last_row < 1:
                    continue

                # Column D: horse names. Pull in one go.
                values = ws.Range(f"D1:D{last_row}").Value
                # values is tuple-of-tuples (or single tuple) depending on range size
                if values is None:
                    continue

                # Pre-normalize meeting keys for faster lookup
                norm_map = {normalize_horse(k): v for k, v in meeting_ratings.items()}

                for idx, row_val in enumerate(values, start=1):
                    cell_val = row_val[0] if isinstance(row_val, (tuple, list)) else row_val
                    if cell_val is None:
                        continue
                    s = str(cell_val).strip()
                    if not s:
                        continue

                    data = norm_map.get(normalize_horse(s))
                    if not data:
                        continue

                    sb_rating = data.get("sb_rating") if isinstance(data, dict) else data
                    win_fixed = data.get("win_fixed") if isinstance(data, dict) else None

                    if sb_rating is not None:
                        ws.Cells(idx, 25).Value = sb_rating  # Y
                    if win_fixed is not None:
                        try:
                            ws.Cells(idx, 22).Value = float(win_fixed)  # V
                        except Exception:
                            ws.Cells(idx, 22).Value = win_fixed

            try:
                wb.Save()
                print(f"Workbook saved via Excel: {excel_file_abs}", flush=True)
            except pywintypes.com_error as ce:
                # Don't crash with traceback; Excel may refuse save if file is opened twice.
                msg = str(ce)
                if "sharing violation" in msg.lower():
                    print(
                        "Excel could not save due to a sharing violation. "
                        "This usually means the workbook is open in another Excel instance/process. "
                        "Close other Excel windows for this file and press Ctrl+S in Excel, or rerun.",
                        flush=True,
                    )
                    return
                print(f"Excel Save() failed: {ce}", flush=True)
                return
        finally:
            try:
                xl.DisplayAlerts = prev_alerts
            except Exception:
                pass

    finally:
        pass


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    except Exception:
        pass
    driver = setup_driver()

    race_links = get_races(driver)

    # If we derived meeting targets from Excel, scrape meeting-by-meeting to avoid name collisions.
    excel_targets = get_excel_targets(FILE_NAME)
    if excel_targets:
        slug_to_meeting = { _slugify(m): m for m in excel_targets.keys() }
        for race_link in race_links:
            parsed = _parse_schedule_href(race_link)
            meeting_slug = parsed[0] if parsed else "RACE"
            meeting_name = slug_to_meeting.get(meeting_slug, meeting_slug)
            extract_sb_rating(driver, race_link, meeting_name)
    else:
        for race_link in race_links:
            extract_sb_rating(driver, race_link, "RACE")

    driver.quit()
    save_sb_to_excel(FILE_NAME, SR)

if __name__ == '__main__':
    main()
